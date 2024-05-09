import telebot
import openpyxl
import os
from telebot import types
from datetime import datetime

# Conexion con el bot
TOKEN = '6871674444:AAGWU6KWGyf13e06-VPKbtdJ2LjENMw1X4I'
bot = telebot.TeleBot(TOKEN)

# Creacion de estados
INICIO, LLEGADA_PLANTA, ESTATUS_CARGA, CUANTAS_UNIDADES, ACTUALIZAR_ESTATINICIO, LLEGADA_PLANTA, ESTATUS_CARGA, CUANTAS_UNIDADES, ACTUALIZAR_ESTATUS, ESTATUS_CARGA_CON_ORDEN, NUEVO_ESTADO, HA_TENIDO_CONTRATIEMPO, SIN_CONTRATIEMPOS, CON_CONTRATIEMPOS, CON_CONTRATIEMPOS_TIPO, CONFIRMAR_LLEGADA_CLIENTE, ESPECIFICAR_CONTRATIEMPO, PREGUNTAR_MUESTRA, PREGUNTAR_BASCULA, CONFIRMACION_BASCULA, CONFIRMACION_RECIBIDO, CONFRIMACION_MUESTRA, DESCARGA_EN_PROGRESO, TERMINAR_DESCARGA, ADJUNTAR_EVIDENCIA, RUMBO_A_LA_PLANTA, OTRA_ENTREGA, LLEGADA_PLANTA, ESPERANDO_NUMERO_CARGA = range(29)


# Variable para almacenar el estado actual de cada usuario
user_states = {}

def verificar_archivo_excel():
    nombre_archivo = 'respuestas_bot.xlsx'
    if not os.path.exists(nombre_archivo):
        wb = openpyxl.Workbook()
        wb.save(nombre_archivo)

# Llamar a la función para verificar el archivo Excel antes de ejecutar el bot
verificar_archivo_excel()

# Función para guardar respuestas en Excel con fecha, hora y nombre del usuario
def guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta):
    # Obtener la fecha y hora actual
    fecha_hora_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Obtener el nombre del usuario o número de quien contestó
    nombre_usuario = bot.get_chat(user_id).first_name
    
    # Abrir el archivo Excel
    wb = openpyxl.load_workbook('respuestas_bot.xlsx')
    ws = wb.active
    
    # Agregar los datos a la siguiente fila vacía
    fila = ws.max_row + 1
    ws.cell(row=fila, column=1, value=fecha_hora_actual)
    ws.cell(row=fila, column=2, value=nombre_usuario)
    ws.cell(row=fila, column=3, value=respuesta)
    #ws.cell(row=fila, column=4, value=chat_id)  # Agregar el ID del grupo
    #ws.cell(row=fila, column=5, value=user_id)  # Agregar el ID del usuario
    
    # Guardar el archivo Excel
    wb.save('respuestas_bot.xlsx')

# Manejadores de comandos
@bot.message_handler(commands=['start'])
def handle_start(message):
    user_id = message.from_user.id
    chat_id = message.chat.id 
    user_states[chat_id] = INICIO  # Inicializamos el estado del usuario
    nombre_usuario = message.from_user.first_name
    respuesta = message.text
    guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    bot.send_message(chat_id, "Buenos dias ¿Desea iniciar el proceso de seguimiento")
    bot.send_message(chat_id, "Desea iniciar proceso?", reply_markup=get_keyboard(chat_id))

# Manejador de mensajes para la opción "Iniciar proceso"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == INICIO)
def handle_inicio(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    user_states[chat_id] = LLEGADA_PLANTA  # Cambiamos el estado del usuario
    respuesta = message.text
    nombre_usuario = message.from_user.first_name
    guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    bot.send_message(chat_id, "Has iniciado el proceso.")
    bot.send_message(chat_id, "Ah llegado a la planta?", reply_markup=get_keyboard(chat_id))

# Manejador de mensajes para el estado "LLEGADA_PLANTA"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == LLEGADA_PLANTA)
def handle_llegada_planta(message):
    user_id = message.from_user.id
    chat_id = message.chat.id 
    nombre_usuario = message.from_user.first_name
    if message.text.lower() == "sí":
        user_states[chat_id] = CUANTAS_UNIDADES
        respuesta = message.text
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Ninguna"),types.KeyboardButton("1"), types.KeyboardButton("2"))
        markup.row(types.KeyboardButton("3"), types.KeyboardButton("Más de tres"))
        markup.row(types.KeyboardButton("No hay producto"))
        bot.send_message(chat_id, "¿Cuántas unidades hay delante de ti?", reply_markup=markup)
    elif message.text.lower() == "no":
        bot.send_message(chat_id, "Está bien, por favor, en diez minutos vuelve a reportarte.")
        user_states[chat_id] = LLEGADA_PLANTA
        bot.send_message(chat_id, "Has llegado a la planta?", reply_markup=get_keyboard(chat_id))
    else:
        bot.send_message(chat_id, "Por favor, selecciona 'Sí' o 'No'.")

# Manejador de mensajes para el estado "CUANTAS_UNIDADES"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == CUANTAS_UNIDADES)
def handle_cuantas_unidades(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    user_states[chat_id] = ACTUALIZAR_ESTATUS
    respuesta = message.text
    nombre_usuario = message.from_user.first_name
    guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    bot.send_message(chat_id, "Muy bien, por favor, actualiza tu estatus.", reply_markup=get_keyboard_actualizar_estatus(chat_id))

# Función para obtener el teclado en función del estado "ACTUALIZAR_ESTATUS"
def get_keyboard_actualizar_estatus(chat_id):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(types.KeyboardButton("Ya entré a cargar"))
    markup.row(types.KeyboardButton("No he entrado a cargar"))
    markup.row(types.KeyboardButton("Sigue sin haber producto"))
    return markup

# Manejador de mensajes para el estado "ACTUALIZAR_ESTATUS"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == ACTUALIZAR_ESTATUS)
def handle_actualizar_estatus(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    nombre_usuario = message.from_user.first_name
    if message.text.lower() == "ya entré a cargar":
        # Cambiar el estado del usuario según lo necesario
        user_states[chat_id] = NUEVO_ESTADO
        respuesta = message.text
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Ya cargué"))
        bot.send_message(chat_id, "Perfecto, favor de avisar cuando termine de cargar.", reply_markup=markup)
        # Puedes agregar más lógica o cambios de estado según tus necesidades
    elif message.text.lower() == "no he entrado a cargar" or message.text.lower() == "sigue sin haber producto":
        bot.send_message(chat_id, "Entendido, favor de notificar su estatus en 15 minutos.")
        # Volver al estado de actualización
        user_states[chat_id] = ACTUALIZAR_ESTATUS
        nombre_usuario = message.from_user.first_name
        respuesta = message.text
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    else:
        bot.send_message(chat_id, "Por favor, seleccione una opción válida.")

# Manejador de mensajes para el estado "NUEVO_ESTADO"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == NUEVO_ESTADO and message.text.lower() == "he iniciado el viaje con el nuevo cliente" or message.text.lower() == "ya cargué")
def handle_rumbo_al_cliente(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    bot.send_message(chat_id, "He iniciado el viaje hacia el cliente.")
    # Cambiar al nuevo estado
    user_states[chat_id] = HA_TENIDO_CONTRATIEMPO
    respuesta = message.text
    nombre_usuario = message.from_user.first_name
    guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    
    # Mostrar los botones "Sin contratiempo" y "Con contratiempo"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(types.KeyboardButton("Sin contratiempo"), types.KeyboardButton("Con contratiempo"))
    bot.send_message(chat_id, "¿Ha tenido algún contratiempo en el camino? Seleccione una opción:", reply_markup=markup)

# Manejador de mensajes para el estado "HA_TENIDO_CONTRATIEMPO"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == HA_TENIDO_CONTRATIEMPO)
def handle_contratiempo(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    if message.text.lower() == "sin contratiempo":
        
        # Cambiar al nuevo estado
        user_states[chat_id] = SIN_CONTRATIEMPOS  # No asignar SIN_CONTRATIEMPOS aquí
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
        
        # Mostrar el botón "Ya llegué con el cliente"
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Voy sin contratiempos con el cliente"))
        bot.send_message(chat_id, "Muchas gracias, favor se seguir actualizando su estado", reply_markup=markup)
    elif message.text.lower() == "con contratiempo":
        bot.send_message(chat_id, "Lamentamos saber eso. ¿Puedes proporcionar más detalles?")
        
        # Cambiar al nuevo estado
        user_states[chat_id] = CON_CONTRATIEMPOS
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
        
        # Llamar a la función para manejar el estado CON_CONTRATIEMPOS
        handle_con_contratiempos(message)
    else:
        bot.send_message(chat_id, "Por favor, selecciona una opción válida.")

# Manejador de mensajes para el estado "SIN_CONTRATIEMPOS"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == SIN_CONTRATIEMPOS)
def handle_ya_llegue(message):
    user_id = message.from_user.id
    chat_id = message.chat.id 
    bot.send_message(chat_id, "Perfecto. Gracias por la actualización.")
    
    # Cambiar al nuevo estado
    user_states[chat_id] = CONFIRMAR_LLEGADA_CLIENTE
    respuesta = message.text
    
    nombre_usuario = message.from_user.first_name
    guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)

    # Mostrar los botones "Si" y "No" para la pregunta sobre la llegada con el cliente
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
    bot.send_message(chat_id, "¿Se logrará llegar con el cliente?", reply_markup=markup)


# Manejador de mensajes para el estado "CON_CONTRATIEMPOS"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == CON_CONTRATIEMPOS)
def handle_con_contratiempos(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    #bot.send_message(chat_id, "Se tiene que especificar el contratiempo")

    # Cambiar al nuevo estado
    user_states[chat_id] = CON_CONTRATIEMPOS_TIPO
    
    # Mostrar los botones "Infraccion", "Accidente", "Tráfico intenso", "Otro"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(types.KeyboardButton("Infracción"), types.KeyboardButton("Accidente"))
    markup.row(types.KeyboardButton("Tráfico intenso"), types.KeyboardButton("Otro"))
    bot.send_message(chat_id, "Selecciona el tipo de contratiempo:", reply_markup=markup)
    

# Manejador de mensajes para el estado "CON_CONTRATIEMPOS_TIPO"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == CON_CONTRATIEMPOS_TIPO)
def handle_con_contratiempos_tipo(message):
    user_id = message.from_user.id
    chat_id = message.chat.id 
    tipo_contratiempo = message.text.lower()
    respuesta = message.text
    nombre_usuario = message.from_user.first_name
    guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)

    if tipo_contratiempo == "otro":
        # Si el contratiempo es "Otro", pide al usuario que lo especifique
        bot.send_message(chat_id, "Especifique el tipo de contratiempo.")
        # Cambia al nuevo estado
        user_states[chat_id] = ESPECIFICAR_CONTRATIEMPO
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    elif tipo_contratiempo != "otro":
        # Si el contratiempo es "Infracción", "Accidente" o "Tráfico intenso", pasa directamente a la confirmación de llegada
        # Cambiar al nuevo estado
        user_states[chat_id] = CONFIRMAR_LLEGADA_CLIENTE

        # Mostrar los botones "Si" y "No" para la pregunta sobre la llegada con el cliente
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        bot.send_message(chat_id, "¿Se logrará llegar con el cliente?", reply_markup=markup)


# Manejador de mensajes para el estado "ESPECIFICAR_CONTRATIEMPO"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == ESPECIFICAR_CONTRATIEMPO)
def handle_especificar_contratiempo(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    contratiempo_especifico = message.text

    # Realiza acciones adicionales según el contratiempo especificado
    bot.send_message(chat_id, f"Contratiempo especificado: {contratiempo_especifico}")

    # Después de manejar el contratiempo específico, pasa a la confirmación de llegada
    # Cambiar al nuevo estado
    user_states[chat_id] = CONFIRMAR_LLEGADA_CLIENTE
    respuesta = message.text
    nombre_usuario = message.from_user.first_name
    guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)

    # Mostrar los botones "Si" y "No" para la pregunta sobre la llegada con el cliente
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
    bot.send_message(chat_id, "¿Se logrará llegar con el cliente?", reply_markup=markup)


# Manejador de mensajes para el estado "CONFIRMAR_LLEGADA_CLIENTE"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == CONFIRMAR_LLEGADA_CLIENTE)
def handle_confirmar_llegada_cliente(message):
    user_id = message.from_user.id
    chat_id = message.chat.id

    # Verificar la respuesta del usuario
    if message.text.lower() == "si":
        # Cambiar al nuevo estado
        user_states[chat_id] = PREGUNTAR_MUESTRA
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)

        # Mostrar los botones "Si" y "No" para la pregunta sobre muestra
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        bot.send_message(chat_id, "¿El cliente sacará muestra?", reply_markup=markup)
        
    elif message.text.lower() == "no":
        bot.send_message(chat_id, "Entendido, se tomaran las medidas necesarias")
        # Puedes terminar el chat o realizar otras acciones según sea necesario
        # Por ejemplo, puedes reiniciar el estado o eliminar el estado del usuario
        del user_states[chat_id]
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
        # También puedes cerrar el teclado al final del chat
        markup = types.ReplyKeyboardRemove(selective=False)
        bot.send_message(chat_id, "¡Hasta luego!", reply_markup=markup)

    else:
        # Si la respuesta no es "si" ni "no", simplemente enviar un mensaje indicando que la respuesta no es válida
        bot.send_message(chat_id, "Por favor, selecciona 'Si' o 'No'.")

# Manejador de mensajes para el estado "PREGUNTAR_MUESTRA"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == PREGUNTAR_MUESTRA)
def handle_preguntar_muestra(message):
    user_id = message.from_user.id
    chat_id = message.chat.id

    # Verificar la respuesta del usuario
    if message.text.lower() == "si":
        bot.send_message(chat_id, "Perfecto. ¿La muestra fue satisfactoria?")
        user_states[chat_id] = CONFRIMACION_MUESTRA
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    elif message.text.lower() == "no":
        # Cambiar al nuevo estado para preguntar sobre la bascula
        user_states[chat_id] = PREGUNTAR_BASCULA
        respuesta = message.text
        chat_id = message.chat.id  # Obtener el ID del grupo
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)

        # Mostrar los botones "Si" y "No" para la pregunta sobre la bascula
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        bot.send_message(chat_id, "¿La carga irá a báscula?", reply_markup=markup)
    else:
        # Si la respuesta no es "si" ni "no", simplemente preguntar de nuevo sobre la muestra
        # sin cambiar de estado
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        bot.send_message(chat_id, "¿El cliente sacará muestra?", reply_markup=markup)

# Manejador de mensajes para el estado "CONFIRMACION_MUESTRA"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == CONFRIMACION_MUESTRA)
def handle_confirmacion_muestra(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    if message.text.lower() == "si":
        # Cambiar al nuevo estado para preguntar sobre la bascula
        user_states[chat_id] = PREGUNTAR_BASCULA
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)

        # Mostrar los botones "Si" y "No" para la pregunta sobre la bascula
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        bot.send_message(chat_id, "¿La carga irá a báscula?", reply_markup=markup)
    elif message.text.lower() == "no":
        bot.send_message(chat_id, "Entendido, el equipo de ventas le dará seguimiento, espera indicaciones")
        bot.send_message(chat_id, "Favor de notificar si ha iniciado un nuevo viaje con el cliente.")
        
        # Mostrar el botón "He iniciado el viaje con el nuevo cliente"
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("He iniciado el viaje con el nuevo cliente"))
        
        # Cambiar al nuevo estado
        user_states[chat_id] = NUEVO_ESTADO
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
        bot.send_message(chat_id, "¿Has iniciado el viaje con el nuevo cliente?", reply_markup=markup)
    else:
        # Si la respuesta no es válida, solicitar una respuesta válida
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        bot.send_message(chat_id, "Por favor, selecciona una opción válida.", reply_markup=markup)

# Manejador de mensajes para el nuevo estado "PREGUNTAR_BASCULA"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == PREGUNTAR_BASCULA)
def handle_preguntar_bascula(message):
    user_id = message.from_user.id
    chat_id = message.chat.id 

    # Verificar la respuesta del usuario
    if message.text.lower() == "si":
        bot.send_message(chat_id, "Favor de compartir comprobante")
        user_states[chat_id] = CONFIRMACION_BASCULA
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
        # Aquí puedes cambiar a otro estado si es necesario
    elif message.text.lower() == "no":
        bot.send_message(chat_id, "Entendido. Favor de notificar cuando le reciban la carga")
        # Cambiar al nuevo estado (por ejemplo, ir a CONFIRMACION_RECIBIDO)
        user_states[chat_id] = CONFIRMACION_RECIBIDO
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    else:
        # Si la respuesta no es "si" ni "no", simplemente preguntar de nuevo sobre la bascula
        # sin cambiar de estado
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        bot.send_message(chat_id, "¿La carga irá a báscula?", reply_markup=markup)

# Manejador de mensajes para el nuevo estado "CONFIRMACION_BASCULA"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == CONFIRMACION_BASCULA, content_types=['text', 'photo'])
def handle_confirmar_bascula(message):
    user_id = message.from_user.id
    chat_id = message.chat.id

    if message.content_type == 'text':
        # Procesar mensaje de texto
        texto = message.text
        # Aquí puedes realizar acciones adicionales con el texto si es necesario

    elif message.content_type == 'photo':
        # Procesar foto
        # Puedes acceder a la foto a través de message.photo
        photo_id = message.photo[-1].file_id
        # Aquí puedes realizar acciones adicionales con la foto si es necesario

    # Enviar el mensaje de confirmación
    bot.send_message(chat_id, "Gracias. Favor de notificar cuando le reciban la carga.")

    # Cambiar al nuevo estado
    user_states[chat_id] = CONFIRMACION_RECIBIDO
    respuesta = message.text
    nombre_usuario = message.from_user.first_name
    guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)

# Manejador de mensajes para el nuevo estado "CONFIRMACION_RECIBIDO"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == CONFIRMACION_RECIBIDO)
def handle_confirmacion_recibido(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    # Verificar la respuesta del usuario
    if message.text.lower() == "ya me están recibiendo":
        # Enviar mensaje de confirmación
        bot.send_message(chat_id, "Perfecto. Favor de informar cuando estés descargando.")
        # Puedes cambiar a otro estado si es necesario
        user_states[chat_id] = DESCARGA_EN_PROGRESO
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
        # Enviar el teclado correspondiente al nuevo estado
        keyboard = get_keyboard(chat_id)
        bot.send_message(chat_id, "¿Cómo va la descarga?", reply_markup=keyboard)
        # Agregar lógica adicional según sea necesario
    elif message.text.lower() == "no me han recibido":
        # Enviar mensaje de confirmación
        bot.send_message(chat_id, "Entendido. Vuelve a mandar tu estado en 15 minutos.")
        # Puedes cambiar a otro estado si es necesario
        user_states[chat_id] = CONFIRMACION_RECIBIDO
        respuesta = message.text
        chat_id = message.chat.id  # Obtener el ID del grupo
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
        # Mostrar botones "Ya me están recibiendo" y "No me han recibido"
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Ya me están recibiendo"), types.KeyboardButton("No me han recibido"))
        bot.send_message(chat_id, "¿Te están recibiendo?", reply_markup=markup)
    else:
        # Si la respuesta no es válida, solicitar una respuesta válida
        keyboard = get_keyboard(chat_id)
        bot.send_message(chat_id, "Por favor, selecciona una opción válida.", reply_markup=keyboard)


# Manejador de mensajes para el nuevo estado "DESCARGA_EN_PROGRESO"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == DESCARGA_EN_PROGRESO)
def handle_descarga_en_progreso(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    # Verificar si el mensaje es "Estoy descargando"
    if message.text.lower() == "estoy descargando":
        bot.send_message(chat_id, "Gracias por la actualización de tu estado.")
        # Cambiar al nuevo estado "TERMINAR_DESCARGA"
        user_states[chat_id] = TERMINAR_DESCARGA
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
        # Mostrar los botones "La descarga ha terminado"
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("La descarga ha terminado"))
        bot.send_message(chat_id, "¿La descarga ha terminado?", reply_markup=markup)

# Manejador de mensajes para el nuevo estado "TERMINAR_DESCARGA"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == TERMINAR_DESCARGA)
def handle_terminar_descarga(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    # Verificar si el mensaje es "La descarga ha terminado"
    if message.text.lower() == "la descarga ha terminado":
        bot.send_message(chat_id, "Favor de adjuntar evidencia.")
        # Cambiar al nuevo estado "ADJUNTAR_EVIDENCIA"
        user_states[chat_id] = ADJUNTAR_EVIDENCIA
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)

# Agrega la nueva función handle_adjuntar_evidencia
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == ADJUNTAR_EVIDENCIA, content_types=['photo'])
def handle_adjuntar_evidencia(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    # Verificar si el mensaje es una foto
    if message.content_type == 'photo':
        # Puedes acceder a la foto a través de message.photo
        photo_id = message.photo[-1].file_id
        # Aquí puedes realizar acciones adicionales con la foto si es necesario
        bot.send_message(chat_id, "¡Gracias por adjuntar la evidencia!")

        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        bot.send_message(chat_id, "¿Tienes otra entrega?", reply_markup=markup)
        # Cambiar al nuevo estado "OTRA_ENTREGA"
        user_states[chat_id] = OTRA_ENTREGA
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    else:
        # Si el mensaje no es una foto, solicitar nuevamente la evidencia
        bot.send_message(chat_id, "Favor de adjuntar una foto como evidencia.")

# Manejador de mensajes para el nuevo estado "RUMBO_A_LA_PLANTA"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == RUMBO_A_LA_PLANTA)
def handle_rumbo_a_la_planta(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    # Verificar si el mensaje es "Rumbo a la planta"
    if message.text.lower() == "rumbo a la planta":
        bot.send_message(chat_id, "Favor de notificar cuando llegue a la planta.")
        # Cambiar al nuevo estado "LLEGADA_PLANTA"
        user_states[chat_id] = LLEGADA_PLANTA
        respuesta = message.text
        nombre_usuario = message.from_user.first_name
        guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)

        # Mostrar el botón "He llegado a la planta"
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.row(types.KeyboardButton("He llegado a la planta"))
        bot.send_message(chat_id, "¿Has llegado a la planta?", reply_markup=markup)

# Manejador de mensajes para el estado "LLEGADA_PLANTA"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == LLEGADA_PLANTA)
def handle_llegada_planta(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    respuesta = message.text
    nombre_usuario = message.from_user.first_name
    guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    # Verificar si el mensaje es "He llegado a la planta"
    if message.text.lower() == "he llegado a la planta":
        bot.send_message(chat_id, "¡Excelente! Gracias por notificar. ¡Hasta luego!")
        del user_states[chat_id]
        # También puedes cerrar el teclado al final del chat
        markup = types.ReplyKeyboardRemove(selective=False)
        bot.send_message(chat_id, "¡Hasta luego!", reply_markup=markup)
    else:
        # Si la respuesta no es válida, solicitar una respuesta válida
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        markup.row(types.KeyboardButton("He llegado a la planta"))


# Manejador de mensajes para el nuevo estado "OTRA_ENTREGA"
@bot.message_handler(func=lambda message: user_states.get(message.chat.id) == OTRA_ENTREGA)
def handle_otra_entrega(message):
    user_id = message.from_user.id
    chat_id = message.chat.id
    respuesta = message.text
    nombre_usuario = message.from_user.first_name
    guardar_respuesta_en_excel(chat_id, user_id, nombre_usuario, respuesta)
    # Verificar la respuesta del usuario
    if message.text.lower() == "si":
        bot.send_message(chat_id, "Favor de adjuntar el número de la siguiente carga y notificar cuando haya iniciado el nuevo viaje con el cliente..")
        
        # Mostrar el botón "He iniciado el viaje con el nuevo cliente"
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("He iniciado el viaje con el nuevo cliente"))
        
        # Cambiar al nuevo estado
        user_states[chat_id] = NUEVO_ESTADO
        bot.send_message(chat_id, "¿Has iniciado el viaje con el nuevo cliente?", reply_markup=markup)
    elif message.text.lower() == "no":
        # Cambiar al nuevo estado "RUMBO_A_LA_PLANTA"
        user_states[chat_id] = RUMBO_A_LA_PLANTA
        # Mostrar el botón "Rumbo a la planta"
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Rumbo a la planta"))
        bot.send_message(chat_id, "¿Listo para ir rumbo a la planta?", reply_markup=markup)
    else:
        # Si la respuesta no es válida, solicitar una respuesta válida
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        bot.send_message(chat_id, "Por favor, selecciona una opción válida.", reply_markup=markup)


# Función para obtener el teclado en función del estado
def get_keyboard(user_id):
    current_state = user_states.get(user_id, INICIO)  # Obtenemos el estado actual del usuario
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)

    if current_state == INICIO:
        keyboard.add(types.KeyboardButton("Iniciar proceso"))
    elif current_state == LLEGADA_PLANTA:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Sí"), types.KeyboardButton("No"))
        return markup
    elif current_state == CON_CONTRATIEMPOS:
        # Agrega los botones necesarios para la pregunta sobre contratiempos
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Con contratiempo"))
        markup.row(types.KeyboardButton("Sin contratiempo"))
        return markup
    elif current_state == SIN_CONTRATIEMPOS:
        # Agrega el botón para "Ya llegué con el cliente"
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Ya llegué con el cliente"))
        return markup
    elif current_state == CONFIRMAR_LLEGADA_CLIENTE:
        # Agrega los botones para la confirmación de llegada
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        return markup
    elif current_state == CON_CONTRATIEMPOS_TIPO:
        # Agrega los botones necesarios para especificar el tipo de contratiempo
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Infracción"), types.KeyboardButton("Accidente"))
        markup.row(types.KeyboardButton("Tráfico intenso"), types.KeyboardButton("Otro"))
        return markup
    elif current_state == PREGUNTAR_BASCULA:
        # Agrega los botones para la pregunta sobre la bascula
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        return markup
    elif current_state == CONFRIMACION_MUESTRA:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        return markup
    elif current_state == CONFIRMACION_RECIBIDO:
        # Agrega los botones para la confirmación recibido
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Ya me están recibiendo"), types.KeyboardButton("No me han recibido"))
        return markup
    elif current_state == DESCARGA_EN_PROGRESO:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Estoy descargando"))
        return markup
    elif current_state == TERMINAR_DESCARGA:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("La descarga ha terminado"))
        return markup
    elif current_state == ADJUNTAR_EVIDENCIA:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        return markup
    elif current_state == RUMBO_A_LA_PLANTA:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Rumbo a la planta"))
        return markup
    elif current_state == OTRA_ENTREGA:
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        markup.row(types.KeyboardButton("Si"), types.KeyboardButton("No"))
        return markup
    else:
        # Puedes manejar otros estados según sea necesario
        bot.send_message(user_id, "Estado no reconocido.")

    return keyboard


if __name__ == "__main__":
    try:
        bot.polling(none_stop=True)
    except Exception as e:
        print(f"Error: {e}")
